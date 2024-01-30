import os
import shutil
os.environ.setdefault("DJANGO_SETTINGS_MODULE", "reshuffle.settings")
import json
import openpyxl
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Font
import re
from datetime import datetime
from random import shuffle, choice, choices
import django
from django.utils.translation import gettext_lazy as _
from django.template.loader import render_to_string
django.setup()
from reshuffle.settings import BASE_DIR, MEDIA_ROOT
from main.models import *
from main.services.docs.minio_client import MinioClient


class UniqueKey:
    """
    ...
    """

    __BASE36 = "0123456789ABCDEFGHIJKLMNOPQRSTUVWXYZ"

    def __init__(self, length: int) -> None:
        self.length = length

    def create(self) -> str:
        return "".join(choices(self.__BASE36, k=self.length))


class GeneratorJSON:
    """
    ...
    """

    __UNIQUE_KEY_LENGTH = 6
    __OUTPUT_JSON = "data.json"

    def __init__(self, sbj_id: int, date: str) -> None:
        self.__data = {
            "subject": {
                "id": sbj_id,
                "title": Subject.objects.get(id=sbj_id).sbj_title,
                "inst_content": Subject.objects.get(id=sbj_id).inst_content
            },
            "date": date,
            "doc_header": DocHeader.objects.filter(is_active=True)[0].content,
            "variants": []
        }

    def __get_unique_keys(self, count: int) -> set[str]:
        uk = UniqueKey(self.__UNIQUE_KEY_LENGTH)
        result = [uk.create() for _ in range(count)]
        while len(result) != len(set(result)):
            result.append(uk.create())
        return set(result)

    def __collect_parts(self) -> list[dict]:
        # init result
        result = []
        # populate result with parts
        for part in Part.objects.filter(subject__id=self.__data["subject"]["id"]):
            # create part's info
            info = {
                "id": part.id,
                "title": str(Part.TITLES[part.title]),
                "answer_type": part.answer_type,
                "task_count": part.task_count,
                "difficulty_total": part.total_difficulty,
                "difficulty_generated": 0,
                "inst_content": part.inst_content
            }
            # calculate part's difficulty distribution
            difficulties = list(DIFFICULTIES.keys())
            distribution = [choice(difficulties) for _ in range(part.task_count)]
            while sum(distribution) != part.total_difficulty:
                if sum(distribution) > part.total_difficulty:
                    i = distribution.index(choice(list(set(difficulties[1:]) & set(distribution))))
                    distribution[i] -= 1
                else:
                    i = distribution.index(choice(list(set(difficulties[:-1]) & set(distribution))))
                    distribution[i] += 1
            shuffle(distribution)
            # create part's material
            material = []
            for position in range(1, part.task_count + 1):
                tasks = Task.objects.filter(part=part, position=position, is_active=True)
                if tasks:
                    # choose task
                    tasks_filtered = tasks.filter(difficulty=distribution[position - 1])
                    task = choice(tasks_filtered) if tasks_filtered else choice(tasks)
                    # choose options
                    options = Option.objects.none()
                    if info["answer_type"] == 0:
                        options = (
                                Option.objects.filter(task=task, is_answer=False).order_by("?")[:3] |
                                Option.objects.filter(task=task, is_answer=True).order_by("?")[:1]
                        ).order_by("?")
                    elif info["answer_type"] == 1:
                        options = Option.objects.filter(task=task, is_answer=True)
                    # update info's difficulty_generated field
                    info["difficulty_generated"] += task.difficulty
                    # append material
                    material.append({
                        "id": task.id,
                        "position": f"{Part.TITLES[part.title]}{position}",
                        "difficulty": task.difficulty,
                        "content": task.content,
                        "options": [{"id": o.id, "content": o.content, "is_answer": o.is_answer} for o in options]
                    })
            # add part's info & material to result
            result.append({"info": info, "material": material})
        # return populated result
        return result

    def generate(self, count: int) -> dict:
        for unique_key in self.__get_unique_keys(count):
            self.__data["variants"].append({"unique_key": unique_key, "parts": self.__collect_parts()})
        return self.__data

    def save(self, path: str) -> None:
        with open(os.path.join(path, self.__OUTPUT_JSON), "w", encoding="UTF-8") as f:
            json.dump(self.__data, f, ensure_ascii=False, indent=2)


class GeneratorXLSX:
    """
    ...
    """

    __BASE_PATH = os.path.join(MEDIA_ROOT, "base", "sheets.xlsx")
    __WS_NAME = "blank"
    __OUTPUT_XLSX = "sheets.xlsx"

    __BORDER_THIN = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    def __init__(self, custom_base_path: str = None) -> None:
        self.__wb_path = custom_base_path if custom_base_path else self.__BASE_PATH
        self.__wb = openpyxl.load_workbook(self.__wb_path)
        self.__ws = self.__wb[self.__WS_NAME]
        self.__sample_created = False

    def __sample(self, sbj_title: str, date: str, doc_header: str, unique_key: str, parts: list) -> None:
        # change sample_created flag
        self.__sample_created = True
        # init labels fields
        self.__ws.title = unique_key
        self.__ws["A1"] = re.sub(re.compile("<.*?>|&([a-z0-9]+|#[0-9]{1,6}|#x[0-9a-f]{1,6});"), "", doc_header)
        self.__ws["A4"] = f"{sbj_title} – " + _("Entrance exams")
        self.__ws["B7"] = str(_("Full name of applicant"))
        self.__ws["B9"] = _("Personnel file") + " №"
        self.__ws["D13"] = unique_key
        self.__ws["B14"] = _("Variant") + " №"
        self.__ws["P13"] = date
        self.__ws["N14"] = str(_("Exam date"))
        self.__ws["Z14"] = str(_("Signature of applicant"))
        self.__ws["A16"] = f"{sbj_title} – " + _("Answer sheet")
        self.__ws["A18"] = _("Variant") + f" № {unique_key}"
        self.__ws["A63"] = str(_("Correcting wrong marks"))
        self.__ws["A67"] = str(_("Results"))
        self.__ws["A70"] = str(_("Total number of\ncorrectly solved problems"))
        self.__ws["T70"] = str(_("Signature of examiner"))
        # split parts for correct render
        parts_splitted = []
        for part in parts:
            while part["task_count"] // Part.CAPACITIES[part["answer_type"]] > 0:
                parts_splitted.append(part.copy())
                parts_splitted[-1]["task_count"] = Part.CAPACITIES[part["answer_type"]]
                part["task_count"] -= Part.CAPACITIES[part["answer_type"]]
            if part["task_count"]:
                parts_splitted.append(part.copy())
        # init answer sheet
        r_init, c_init = 19, 1
        accum = 0
        for count, part in enumerate(parts_splitted):
            # update accum
            if count and part["title"] == parts_splitted[count - 1]["title"]:
                accum += Part.CAPACITIES[part["answer_type"]]
            else:
                accum = 0
            # render part [type 0: 15]
            if part["answer_type"] == 0:
                # answer nums
                for i in range(4):
                    for j in (1, 33):
                        self.__ws.cell(
                            row=r_init + 3 + 2 * i,
                            column=c_init + j,
                            value=str((i + 1))
                        ).font = Font(name="Gilroy", size=8, bold=True)
                # answer titles & cells
                for i in range(part["task_count"]):
                    self.__ws.cell(
                        row=r_init + 1,
                        column=c_init + 3 + 2 * i,
                        value=f"{part['title']}{i + 1 + accum}"
                    ).font = Font(name="Gilroy", size=8, bold=True)
                    for j in range(4):
                        self.__ws.cell(
                            row=r_init + 3 + 2 * j,
                            column=c_init + 3 + 2 * i
                        ).border = self.__BORDER_THIN
            # render part [type 1: 10]
            elif part["answer_type"] == 1:
                # answer titles & cells
                for i in range(part["task_count"]):
                    self.__ws.cell(
                        row=r_init + 1 + 2 * (i // 2),
                        column=c_init + 1 + (32 * (i % 2)),
                        value=f"{part['title']}{i + 1 + accum}"
                    ).font = Font(name="Gilroy", size=8, bold=True)
                    for j in range(13):
                        self.__ws.cell(
                            row=r_init + 1 + 2 * (i // 2),
                            column=c_init + 3 + (16 * (i % 2)) + j
                        ).border = self.__BORDER_THIN
            # move render window
            r_init += 11

    def __reproduce(self, unique_key: str) -> None:
        # copy sample & change title
        self.__wb.copy_worksheet(self.__ws).title = unique_key
        # change labels fields
        self.__wb[unique_key]["D13"] = unique_key
        self.__wb[unique_key]["A18"] = _("Variant") + f" № {unique_key}"

    def generate(self, data: dict) -> None:
        for variant in data["variants"]:
            if not self.__sample_created:
                self.__sample(
                    data["subject"]["title"],
                    data["date"],
                    data["doc_header"],
                    variant["unique_key"],
                    [{
                        "title": p["info"]["title"],
                        "answer_type": p["info"]["answer_type"],
                        "task_count": p["info"]["task_count"]
                    } for p in variant["parts"]]
                )
            else:
                self.__reproduce(variant["unique_key"])

    def save(self, path: str) -> None:
        path = os.path.join(path, self.__OUTPUT_XLSX)
        if path != self.__wb_path:
            self.__wb.save(path)
        else:
            raise ValueError(
                f"The name of the destination file ({path}) "
                f"cannot be the same as the name of the source file ({self.__wb_path})."
            )


class GeneratorPDF:
    """
    ...
    """

    __TEMPLATE_TASKS_PATH = "docs/template_tasks.html"
    __TEMPLATE_ANSWERS_PATH = "docs/template_answers.html"
    __OUTPUT_TASKS_HTML = "tasks.html"
    __OUTPUT_ANSWERS_HTML = "answers.html"

    def __init__(self) -> None:
        self.__html_tasks = ""
        self.__html_answers = ""

    def generate(self, data: dict) -> None:
        # generate tasks [HTML]
        self.__html_tasks = render_to_string(
            template_name=self.__TEMPLATE_TASKS_PATH,
            context=data | {"base_dir": json.dumps(str(BASE_DIR))}
        )
        # generate answers [HTML]
        self.__html_answers = render_to_string(template_name=self.__TEMPLATE_ANSWERS_PATH, context=data)

    def save(self, path: str) -> None:
        save_list = [(self.__html_tasks, self.__OUTPUT_TASKS_HTML), (self.__html_answers, self.__OUTPUT_ANSWERS_HTML)]
        for content, output in save_list:
            with open(os.path.join(path, output), "w", encoding="UTF-8") as f:
                f.write(content)


class DocumentPackager:
    """
    ...
    """

    __OUTPUT_PATH = os.path.join(MEDIA_ROOT, "docs")
    ARCHIVE_FORMAT = "zip"

    def __init__(self) -> None:
        pass

    def __create_folder(self, name: str) -> str:
        folder = os.path.join(self.__OUTPUT_PATH, name)
        if not os.path.exists(folder):
            os.makedirs(folder)
        return folder

    def __archive_folder(self, folder: str, delete: bool = True) -> str:
        shutil.make_archive(folder, self.ARCHIVE_FORMAT, folder)
        if delete:
            shutil.rmtree(folder)
        return f"{folder}.{self.ARCHIVE_FORMAT}"

    def pack(self, sbj_id: int, count: int, date: str) -> None:
        # create output folder
        folder = self.__create_folder(f"[{datetime.today().strftime('%d-%m-%Y_%H-%M-%S-%f')}][{sbj_id}]")
        # create data [JSON]
        gen_json = GeneratorJSON(sbj_id, date)
        data = gen_json.generate(count)
        gen_json.save(folder)
        # create sheets [XLSX]
        gen_xlsx = GeneratorXLSX()
        gen_xlsx.generate(data)
        gen_xlsx.save(folder)
        # create tasks & answers [PDF]
        gen_pdf = GeneratorPDF()
        gen_pdf.generate(data)
        gen_pdf.save(folder)
        # init object storage client
        mc = MinioClient()
        # send output folder to object storage
        mc.upload_folder(folder)
        # archive & delete output folder
        archive = self.__archive_folder(folder)
        # send archive to object storage
        mc.upload_file(archive)
        # delete archive
        os.remove(archive)
        # TODO: add entry to db
        # ...


if __name__ == "__main__":
    from time import time

    t = time()

    n = 2
    dp = DocumentPackager()
    dp.pack(sbj_id=2, count=n, date="30.01.2024")

    print(time() - t)
    print((time() - t) / n)
